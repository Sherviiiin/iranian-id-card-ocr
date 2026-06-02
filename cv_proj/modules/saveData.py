from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import cv2
import os

def save_data(
    xlsx_path: str,
    headers: list[str],
    values: list,
    sheet_name: str = "Data",
    image= None,
    folder_name="saved_images",
    ext=".jpg"
):
    if len(headers) != 6:
        raise ValueError("headers باید دقیقا 6 تا باشد.")
    if len(values) != 6:
        raise ValueError("values باید دقیقا 6 تا باشد.")

    # کلید یکتا: 3 ستون اول
    key_new = tuple(values[:3])

    # اگر فایل وجود ندارد: بساز + هدر + اولین ردیف داده
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        ws.append(headers)
        ws.append(values)

        # (اختیاری) عرض ستون‌ها
        for c in range(1, 7):
            ws.column_dimensions[get_column_letter(c)].width = 18

        wb.save(xlsx_path)
        #return True, 2

    # فایل وجود دارد: باز کن
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    existing_keys = {}
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value

        # اگر ردیف خالی است، ردش کن
        if (a in (None, "") and b in (None, "") and c in (None, "")):
            continue

        k = (a, b, c)
        if k not in existing_keys:
            existing_keys[k] = r

    # اگر تکراری بود: اضافه نکن
    if not(key_new in existing_keys):
        #return False, existing_keys[key_new]

        # پیدا کردن اولین ردیف کاملاً خالی در 6 ستون اول
        r = 2
        while True:
            if all(ws.cell(row=r, column=c).value in (None, "") for c in range(1, 7)):
                break
            r += 1

        # نوشتن 6 مقدار
        for c, val in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=val)

        wb.save(xlsx_path)

    # اگر پوشه وجود نداشت، بساز
    os.makedirs(folder_name, exist_ok=True)

    # نام یکتا بساز
    filename = values[0]

    # مسیر کامل فایل
    file_path = os.path.join(folder_name, filename + ext)

    if os.path.exists(file_path):
        return False
    
    if image is not None:
        filename = values[0]
    # ذخیره تصویر
    success = cv2.imwrite(file_path, image)
    if not success:
        raise IOError("خطا در ذخیره تصویر")
    
    return True, r
